#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
飞书考勤数据分析工具
从飞书考勤统计 API 拉取全员打卡数据，生成排行榜、异常预警、部门对比等报表。

用法:
    python analyze_attendance.py                    # 默认近30天
    python analyze_attendance.py --days 7           # 近7天
    python analyze_attendance.py --start 20260501 --end 20260531
    python analyze_attendance.py --excel-only       # 仅输出Excel
    python analyze_attendance.py --no-html          # 不生成 HTML 报表
    python analyze_attendance.py --out-dir ./report # 自定义输出目录
"""

import json
import subprocess
import os
import sys
import argparse
from datetime import datetime, timedelta
from collections import defaultdict
from pathlib import Path

# Windows GBK 控制台直接 print 中文会乱码，统一切到 UTF-8
try:
    sys.stdout.reconfigure(encoding='utf-8')
except (AttributeError, ValueError):
    pass

# MSYS(Git Bash) 会把 /open-apis/... 篡改成绝对路径导致 404，关掉路径转换
os.environ['MSYS_NO_PATHCONV'] = '1'
os.environ['MSYS2_ARG_CONV_EXCL'] = '*'


def run_lark_api(path, params, data, as_bot=True):
    """调用 lark-cli api 裸调"""
    params_file = '_tmp_params.json'
    data_file = '_tmp_data.json'

    with open(params_file, 'w', encoding='utf-8') as f:
        json.dump(params, f, ensure_ascii=False)
    with open(data_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)

    identity = '--as bot' if as_bot else ''
    cmd = f'lark-cli api POST {path} --params @{params_file} --data @{data_file} {identity} --json'
    result = subprocess.run(cmd, capture_output=True, shell=True)

    # 清理临时文件
    for f in [params_file, data_file]:
        try:
            os.remove(f)
        except:
            pass

    stdout = result.stdout.decode('utf-8', errors='replace')
    try:
        return json.loads(stdout)
    except json.JSONDecodeError:
        return None


def load_roster(roster_path=None):
    """加载花名册数据"""
    if roster_path is None:
        # 尝试常见路径
        candidates = [
            './roster/roster.json',
            '../roster/roster.json',
            os.path.expanduser('~/roster/roster.json'),
        ]
        for p in candidates:
            if os.path.exists(p):
                roster_path = p
                break

    if roster_path is None or not os.path.exists(roster_path):
        print("错误: 未找到花名册数据。请先运行 feishu-roster skill 拉取花名册。")
        sys.exit(1)

    with open(roster_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    roster = {}
    for row in data.get('rows', []):
        emp_no = row.get('工号', '')
        if emp_no:
            roster[emp_no] = {
                'name': row.get('姓名', ''),
                'department': row.get('部门', ''),
                'status': row.get('在职状态', ''),
                'type': row.get('人员类型', ''),
            }
    return roster


def query_attendance_stats(user_ids, start_date, end_date):
    """批量查询考勤统计数据"""
    path = '/open-apis/attendance/v1/user_stats_datas/query'
    params = {'employee_type': 'employee_no'}
    data = {
        'locale': 'zh',
        'stats_type': 'daily',
        'start_date': start_date,
        'end_date': end_date,
        'user_ids': user_ids,
        'user_id': user_ids[0] if user_ids else ''
    }

    resp = run_lark_api(path, params, data, as_bot=True)

    if resp and resp.get('code') == 0:
        return resp['data'].get('user_datas', []), resp['data'].get('invalid_user_list', [])
    else:
        code = resp.get('code') if resp else 'N/A'
        msg = resp.get('msg', '')[:100] if resp else 'No response'
        print(f"  API 错误: code={code}, msg={msg}")
        return [], []


def extract_checkout_times(user_data):
    """从考勤统计中提取下班打卡时间"""
    user_id = user_data.get('user_id', '')
    name = None
    checkout_times = []

    for data_item in user_data.get('datas', []):
        code = data_item.get('code', '')
        features = data_item.get('features', [])

        for feat in features:
            if feat.get('key') == 'Name':
                name = feat.get('value')

        # 51503-1-2 = 下午/下班打卡记录
        if code == '51503-1-2':
            for feat in features:
                if feat.get('key') == 'PunchTime':
                    punch_time = feat.get('value')
                    if punch_time and punch_time not in ('-', ''):
                        checkout_times.append(punch_time)

    return user_id, name, checkout_times


def time_to_minutes(time_str):
    """HH:MM -> 分钟数"""
    try:
        parts = time_str.split(':')
        if len(parts) == 2:
            return int(parts[0]) * 60 + int(parts[1])
    except:
        pass
    return None


def minutes_to_time(minutes):
    """分钟数 -> HH:MM"""
    hours = int(minutes // 60)
    mins = int(minutes % 60)
    return f"{hours:02d}:{mins:02d}"


def detect_anomalies(stats):
    """检测考勤异常"""
    anomalies = []

    for s in stats:
        times_min = [time_to_minutes(t) for t in s['all_times']]
        times_min = [t for t in times_min if t is not None]

        if not times_min:
            continue

        # 深夜打卡 (>= 21:00)
        late_night = [t for t in times_min if t >= 21 * 60]
        if late_night:
            anomalies.append({
                'name': s['name'],
                'user_id': s['user_id'],
                'type': '深夜打卡',
                'detail': f"{len(late_night)} 次，最晚 {minutes_to_time(max(late_night))}",
                'severity': 'high' if len(late_night) >= 3 else 'medium'
            })

        # 连续加班 (连续3天以上 >= 19:30)
        consecutive = 0
        max_consecutive = 0
        for t in s['all_times']:
            m = time_to_minutes(t)
            if m is not None and m >= 19 * 60 + 30:
                consecutive += 1
                max_consecutive = max(max_consecutive, consecutive)
            else:
                consecutive = 0
        if max_consecutive >= 3:
            anomalies.append({
                'name': s['name'],
                'user_id': s['user_id'],
                'type': '连续加班',
                'detail': f"连续 {max_consecutive} 天下班 ≥ 19:30",
                'severity': 'high' if max_consecutive >= 5 else 'medium'
            })

        # 作息异常 (标准差 > 60 分钟)
        if len(times_min) >= 5:
            avg = sum(times_min) / len(times_min)
            variance = sum((t - avg) ** 2 for t in times_min) / len(times_min)
            std_dev = variance ** 0.5
            if std_dev > 60:
                anomalies.append({
                    'name': s['name'],
                    'user_id': s['user_id'],
                    'type': '作息异常',
                    'detail': f"下班时间波动大 (标准差 {std_dev:.0f} 分钟)",
                    'severity': 'low'
                })

    return anomalies


def write_excel(stats, anomalies, dept_stats, out_path):
    """写入 Excel 报表"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("警告: 未安装 openpyxl，跳过 Excel 输出。运行 pip install openpyxl 安装。")
        return

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    high_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    medium_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def style_header(ws, cols):
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 30)

    # Sheet 1: 平均下班排行
    ws1 = wb.active
    ws1.title = '平均下班排行'
    cols1 = ['排名', '姓名', '工号', '部门', '平均下班', '打卡天数', '最早', '最晚', '标准差(分钟)']
    style_header(ws1, cols1)

    for i, s in enumerate(stats, 1):
        times_min = [time_to_minutes(t) for t in s['all_times']]
        times_min = [t for t in times_min if t is not None]
        std_dev = 0
        if len(times_min) >= 2:
            avg = sum(times_min) / len(times_min)
            variance = sum((t - avg) ** 2 for t in times_min) / len(times_min)
            std_dev = variance ** 0.5

        ws1.append([
            i, s['name'], s['user_id'], s.get('department', ''),
            s['avg_checkout'], s['days'], s['earliest'], s['latest'],
            round(std_dev, 1)
        ])

    auto_width(ws1)

    # Sheet 2: 异常预警
    ws2 = wb.create_sheet('异常预警')
    cols2 = ['姓名', '工号', '部门', '异常类型', '详情', '严重程度']
    style_header(ws2, cols2)

    for a in sorted(anomalies, key=lambda x: {'high': 0, 'medium': 1, 'low': 2}.get(x['severity'], 3)):
        dept = ''
        for s in stats:
            if s['user_id'] == a['user_id']:
                dept = s.get('department', '')
                break
        row_idx = ws2.max_row + 1
        ws2.append([a['name'], a['user_id'], dept, a['type'], a['detail'],
                     {'high': '高', 'medium': '中', 'low': '低'}.get(a['severity'], '')])
        if a['severity'] == 'high':
            for col in range(1, 7):
                ws2.cell(row=row_idx, column=col).fill = high_fill
        elif a['severity'] == 'medium':
            for col in range(1, 7):
                ws2.cell(row=row_idx, column=col).fill = medium_fill

    auto_width(ws2)

    # Sheet 3: 部门对比
    ws3 = wb.create_sheet('部门对比')
    cols3 = ['部门', '人数', '平均下班', '最早平均', '最晚平均', '深夜打卡次数']
    style_header(ws3, cols3)

    for dept_name, ds in sorted(dept_stats.items(), key=lambda x: x[1]['avg_minutes'], reverse=True):
        ws3.append([
            dept_name, ds['count'], minutes_to_time(ds['avg_minutes']),
            minutes_to_time(ds.get('earliest_avg', 0)),
            minutes_to_time(ds.get('latest_avg', 0)),
            ds.get('late_night_count', 0)
        ])

    auto_width(ws3)

    # Sheet 4: 打卡明细
    ws4 = wb.create_sheet('打卡明细')
    cols4 = ['姓名', '工号'] + [f'Day{i+1}' for i in range(len(stats[0]['all_times']) if stats else 0)]
    style_header(ws4, cols4[:2 + 31])  # 限制列数

    for s in stats:
        row = [s['name'], s['user_id']] + s['all_times']
        ws4.append(row[:33])  # 最多31天+2列

    auto_width(ws4)

    wb.save(out_path)
    print(f"Excel 报表已保存: {out_path}")


def _html_escape(text):
    return (str(text).replace('&', '&amp;').replace('<', '&lt;')
            .replace('>', '&gt;').replace('"', '&quot;'))


HTML_CSS = """
* { margin: 0; padding: 0; box-sizing: border-box; }
body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Microsoft YaHei", sans-serif;
  background: #f5f6f8; color: #1f2329; line-height: 1.5; padding: 32px 16px; }
.container { max-width: 960px; margin: 0 auto; }
h1 { font-size: 24px; font-weight: 700; }
h2 { font-size: 18px; font-weight: 600; margin: 32px 0 16px; padding-left: 10px;
  border-left: 4px solid #3370ff; }
.subtitle { color: #8f959e; font-size: 13px; margin-top: 4px; }
.cards { display: flex; gap: 16px; margin-top: 24px; flex-wrap: wrap; }
.card { flex: 1; min-width: 140px; background: #fff; border-radius: 10px; padding: 18px 20px;
  box-shadow: 0 1px 4px rgba(0,0,0,.06); }
.card .num { font-size: 28px; font-weight: 700; color: #3370ff; }
.card .label { font-size: 13px; color: #646a73; margin-top: 4px; }
.panel { background: #fff; border-radius: 10px; padding: 8px 0; box-shadow: 0 1px 4px rgba(0,0,0,.06);
  overflow: hidden; }
table { width: 100%; border-collapse: collapse; font-size: 14px; }
th, td { padding: 10px 14px; text-align: left; }
thead th { color: #8f959e; font-weight: 500; font-size: 12px; border-bottom: 1px solid #e5e6eb; }
tbody tr { border-bottom: 1px solid #f0f1f3; }
tbody tr:last-child { border-bottom: none; }
.rank { width: 28px; height: 28px; line-height: 28px; text-align: center; border-radius: 50%;
  display: inline-block; font-size: 12px; font-weight: 600; background: #f0f1f3; color: #646a73; }
.rank.top1 { background: #fde68a; color: #92400e; }
.rank.top2 { background: #e5e7eb; color: #4b5563; }
.rank.top3 { background: #fed7aa; color: #9a3412; }
.bar-wrap { background: #f0f1f3; border-radius: 4px; height: 8px; width: 100%; min-width: 80px; }
.bar { height: 8px; border-radius: 4px; background: linear-gradient(90deg, #5b8def, #3370ff); }
.time { font-variant-numeric: tabular-nums; font-weight: 600; }
.muted { color: #8f959e; font-variant-numeric: tabular-nums; }
.anomaly { display: flex; align-items: center; gap: 12px; padding: 12px 16px;
  border-bottom: 1px solid #f0f1f3; }
.anomaly:last-child { border-bottom: none; }
.tag { font-size: 11px; font-weight: 600; padding: 2px 8px; border-radius: 4px; white-space: nowrap; }
.tag.high { background: #fde2e2; color: #d83931; }
.tag.medium { background: #fef0d2; color: #ad6800; }
.tag.low { background: #eaeaea; color: #646a73; }
.anomaly .who { font-weight: 600; min-width: 80px; }
.anomaly .what { color: #646a73; font-size: 13px; }
.empty { padding: 24px; text-align: center; color: #8f959e; }
footer { margin-top: 32px; text-align: center; color: #c9cdd4; font-size: 12px; }
"""


def write_html(stats, anomalies, dept_stats, start_date, end_date,
               total_emp, with_data, out_path):
    """生成自包含可视化 HTML 报表"""
    sev_label = {'high': '高', 'medium': '中', 'low': '低'}

    rank_min = min((s['avg_minutes'] for s in stats), default=0)
    rank_max = max((s['avg_minutes'] for s in stats), default=1)
    rank_span = max(rank_max - rank_min, 1)

    parts = []
    parts.append('<!DOCTYPE html><html lang="zh-CN"><head>')
    parts.append('<meta charset="utf-8">')
    parts.append('<meta name="viewport" content="width=device-width, initial-scale=1">')
    parts.append(f'<title>考勤分析报告 {start_date}~{end_date}</title>')
    parts.append(f'<style>{HTML_CSS}</style></head><body><div class="container">')

    # 头部 + 概览卡片
    parts.append('<h1>全公司考勤分析报告</h1>')
    parts.append(f'<div class="subtitle">统计区间 {start_date} ~ {end_date}'
                 f' · 生成于 {datetime.now().strftime("%Y-%m-%d %H:%M")}</div>')
    parts.append('<div class="cards">')
    parts.append(f'<div class="card"><div class="num">{total_emp}</div>'
                 f'<div class="label">在职员工</div></div>')
    parts.append(f'<div class="card"><div class="num">{with_data}</div>'
                 f'<div class="label">有考勤数据</div></div>')
    parts.append(f'<div class="card"><div class="num">{len(anomalies)}</div>'
                 f'<div class="label">异常预警</div></div>')
    parts.append('</div>')

    # 平均下班排行
    parts.append('<h2>平均下班时间排行</h2><div class="panel"><table>')
    parts.append('<thead><tr><th>排名</th><th>姓名</th><th>部门</th>'
                 '<th>平均下班</th><th style="width:30%">相对</th>'
                 '<th>天数</th><th>最早</th><th>最晚</th></tr></thead><tbody>')
    for i, s in enumerate(stats, 1):
        rank_cls = f'top{i}' if i <= 3 else ''
        width = 8 + (s['avg_minutes'] - rank_min) / rank_span * 92
        parts.append(
            f'<tr><td><span class="rank {rank_cls}">{i}</span></td>'
            f'<td>{_html_escape(s["name"])}</td>'
            f'<td class="muted">{_html_escape(s.get("department", ""))}</td>'
            f'<td class="time">{s["avg_checkout"]}</td>'
            f'<td><div class="bar-wrap"><div class="bar" style="width:{width:.0f}%"></div></div></td>'
            f'<td class="muted">{s["days"]}</td>'
            f'<td class="muted">{s["earliest"]}</td>'
            f'<td class="muted">{s["latest"]}</td></tr>')
    parts.append('</tbody></table></div>')

    # 异常预警
    parts.append('<h2>异常预警</h2><div class="panel">')
    if anomalies:
        order = {'high': 0, 'medium': 1, 'low': 2}
        for a in sorted(anomalies, key=lambda x: order.get(x['severity'], 3)):
            sev = a['severity']
            parts.append(
                f'<div class="anomaly"><span class="tag {sev}">{sev_label.get(sev, "")}</span>'
                f'<span class="who">{_html_escape(a["name"])}</span>'
                f'<span class="tag low">{_html_escape(a["type"])}</span>'
                f'<span class="what">{_html_escape(a["detail"])}</span></div>')
    else:
        parts.append('<div class="empty">无异常</div>')
    parts.append('</div>')

    # 部门对比
    parts.append('<h2>部门对比</h2><div class="panel"><table>')
    parts.append('<thead><tr><th>部门</th><th>人数</th><th>平均下班</th>'
                 '<th style="width:35%">对比</th><th>深夜打卡</th></tr></thead><tbody>')
    dept_sorted = sorted(dept_stats.items(), key=lambda x: x[1]['avg_minutes'], reverse=True)
    dept_min = min((d['avg_minutes'] for _, d in dept_sorted), default=0)
    dept_max = max((d['avg_minutes'] for _, d in dept_sorted), default=1)
    dept_span = max(dept_max - dept_min, 1)
    for dept_name, ds in dept_sorted:
        width = 8 + (ds['avg_minutes'] - dept_min) / dept_span * 92
        parts.append(
            f'<tr><td>{_html_escape(dept_name)}</td>'
            f'<td class="muted">{ds["count"]}</td>'
            f'<td class="time">{minutes_to_time(ds["avg_minutes"])}</td>'
            f'<td><div class="bar-wrap"><div class="bar" style="width:{width:.0f}%"></div></div></td>'
            f'<td class="muted">{ds.get("late_night_count", 0)}</td></tr>')
    parts.append('</tbody></table></div>')

    parts.append('<footer>由 feishu-attendance-analyzer 生成</footer>')
    parts.append('</div></body></html>')

    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(''.join(parts))
    print(f"HTML 报表已保存: {out_path}")


def main():
    parser = argparse.ArgumentParser(description='飞书考勤数据分析')
    parser.add_argument('--days', type=int, default=30, help='查询天数（默认30天）')
    parser.add_argument('--start', type=str, help='开始日期 YYYYMMDD')
    parser.add_argument('--end', type=str, help='结束日期 YYYYMMDD')
    parser.add_argument('--excel-only', action='store_true', help='仅输出Excel（不打印终端）')
    parser.add_argument('--no-html', action='store_true', help='不生成 HTML 报表')
    parser.add_argument('--out-dir', type=str, default='.', help='输出目录')
    parser.add_argument('--roster', type=str, help='花名册路径')
    args = parser.parse_args()

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # 计算日期范围
    if args.start and args.end:
        start_date = int(args.start)
        end_date = int(args.end)
    else:
        today = datetime.now()
        start = today - timedelta(days=args.days)
        start_date = int(start.strftime('%Y%m%d'))
        end_date = int(today.strftime('%Y%m%d'))

    print(f"查询范围: {start_date} ~ {end_date}")

    # 加载花名册
    print("加载花名册...")
    roster = load_roster(args.roster)
    emp_ids = [eid for eid, info in roster.items() if info.get('status') == '在职']
    print(f"在职员工: {len(emp_ids)} 人")

    # 分批查询考勤统计
    print("查询考勤统计...")
    batch_size = 200
    batches = [emp_ids[i:i+batch_size] for i in range(0, len(emp_ids), batch_size)]
    all_user_data = []

    for idx, batch in enumerate(batches):
        print(f"  第 {idx+1}/{len(batches)} 批 ({len(batch)} 人)...")
        user_data, invalid = query_attendance_stats(batch, start_date, end_date)
        all_user_data.extend(user_data)
        if invalid:
            print(f"  无权限: {len(invalid)} 人")

    print(f"获取到 {len(all_user_data)} 人的考勤数据")

    # 聚合分析
    user_checkout = defaultdict(list)
    user_names = {}

    for ud in all_user_data:
        uid, name, times = extract_checkout_times(ud)
        if not uid or not times:
            continue
        if name:
            user_names[uid] = name
        for t in times:
            m = time_to_minutes(t)
            if m is not None:
                user_checkout[uid].append((t, m))

    stats = []
    for uid, entries in user_checkout.items():
        name = user_names.get(uid, uid)
        minutes_list = [m for _, m in entries]
        time_strs = [t for t, _ in entries]
        avg_min = sum(minutes_list) / len(minutes_list)

        dept = roster.get(uid, {}).get('department', '')

        stats.append({
            'name': name,
            'user_id': uid,
            'department': dept,
            'avg_checkout': minutes_to_time(avg_min),
            'avg_minutes': avg_min,
            'days': len(minutes_list),
            'earliest': minutes_to_time(min(minutes_list)),
            'latest': minutes_to_time(max(minutes_list)),
            'all_times': time_strs,
        })

    stats.sort(key=lambda x: x['avg_minutes'], reverse=True)

    # 部门统计
    dept_data = defaultdict(lambda: {'minutes': [], 'late_night': 0})
    for s in stats:
        dept = s['department'] or '未分配'
        dept_data[dept]['minutes'].append(s['avg_minutes'])
        late_night = sum(1 for t in s['all_times'] if time_to_minutes(t) and time_to_minutes(t) >= 21 * 60)
        dept_data[dept]['late_night'] += late_night

    dept_stats = {}
    for dept_name, dd in dept_data.items():
        if dd['minutes']:
            avg = sum(dd['minutes']) / len(dd['minutes'])
            dept_stats[dept_name] = {
                'count': len(dd['minutes']),
                'avg_minutes': avg,
                'earliest_avg': min(dd['minutes']),
                'latest_avg': max(dd['minutes']),
                'late_night_count': dd['late_night'],
            }

    # 异常检测
    anomalies = detect_anomalies(stats)

    # 终端输出
    if not args.excel_only:
        print("\n" + "=" * 65)
        print(f"全公司平均下班时间排行榜 ({start_date} ~ {end_date})")
        print("=" * 65)
        print(f"{'排名':<4}  {'姓名':<10}  {'平均下班':<8}  {'天数':<6}  {'最早':<6}  {'最晚':<6}")
        print("-" * 65)

        for i, s in enumerate(stats[:20], 1):
            print(f"{i:<4}  {s['name']:<10}  {s['avg_checkout']:<8}  {s['days']:<6}  {s['earliest']:<6}  {s['latest']:<6}")

        if len(stats) > 20:
            print(f"\n... 共 {len(stats)} 人有考勤数据")

        if anomalies:
            print("\n" + "=" * 65)
            print("异常预警")
            print("=" * 65)
            severity_icon = {'high': '[!]', 'medium': '[~]', 'low': '[-]'}
            for a in sorted(anomalies, key=lambda x: {'high': 0, 'medium': 1, 'low': 2}.get(x['severity'], 3)):
                icon = severity_icon.get(a['severity'], '')
                print(f"  {icon} {a['name']}: {a['type']} - {a['detail']}")

    # Excel 输出
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    excel_path = out_dir / f"考勤分析_{ts}.xlsx"
    write_excel(stats, anomalies, dept_stats, str(excel_path))

    # HTML 输出
    if not args.no_html:
        html_path = out_dir / f"考勤分析_{ts}.html"
        write_html(stats, anomalies, dept_stats, start_date, end_date,
                   len(emp_ids), len(stats), str(html_path))

    # 保存 JSON
    json_path = out_dir / 'attendance_analysis.json'
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)
    print(f"JSON 数据已保存: {json_path}")


if __name__ == '__main__':
    main()
