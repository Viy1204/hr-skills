# -*- coding: utf-8 -*-
"""
从飞书人事(标准版) ehr/v1/employees 拉取最新在职花名册，映射部门名，导出 Excel + JSON。

为什么这么调（四个坑，都已踩平）：
- ehr 接口用 tenant_access_token，必须 `--as bot`；通讯录部门树走 user。
- Windows 的 Git Bash(MSYS) 会把 `/open-apis/...` 篡改成绝对路径导致 404，
  所以子进程环境里强制 MSYS_NO_PATHCONV=1 / MSYS2_ARG_CONV_EXCL=*。
- lark-cli 的 `--params @file` 只认正斜杠路径，反斜杠会被判 invalid file path，
  所以参数临时文件写在 cwd 下、用纯文件名引用。
- `--page-all` 的进度行会混进 stdout，解析时从第一个 '{' 截断。

用法:
  python pull_roster.py --out-dir ./roster [--status 2] [--xlsx-name 花名册.xlsx]
"""
import argparse, json, os, subprocess, sys, datetime

sys.stdout.reconfigure(encoding="utf-8")

ENV = dict(os.environ)
ENV["MSYS_NO_PATHCONV"] = "1"
ENV["MSYS2_ARG_CONV_EXCL"] = "*"

EMP_TYPE = {1: "正式", 2: "实习", 3: "顾问", 4: "外包", 5: "临时"}
STATUS = {1: "待入职", 2: "在职", 3: "取消入职", 4: "待离职", 5: "离职"}

_seq = [0]


def api(method, path, params, ident="bot", page_all=False, workdir="."):
    """裸调 lark-cli api，返回解析后的 JSON（失败返回带 _err 的 dict）。"""
    _seq[0] += 1
    # lark-cli 的 --params @file 只认「cwd 内的相对路径」，所以参数文件写进 workdir，
    # 子进程 cwd 设为 workdir，引用时用纯文件名。
    fname = f"_p{_seq[0]}.json"
    fp = os.path.join(workdir, fname)
    json.dump(params or {}, open(fp, "w", encoding="utf-8"))
    cmd = f'lark-cli api {method} {path} --params @{fname} --as {ident} --format json'
    if page_all:
        cmd += " --page-all"
    r = subprocess.run(cmd, shell=True, capture_output=True, env=ENV, cwd=workdir)
    try:
        os.remove(fp)
    except OSError:
        pass
    out = r.stdout.decode("utf-8", "replace")
    i = out.find("{")
    if i < 0:
        return {"_err": r.stderr.decode("utf-8", "replace")[:400] or out[:400]}
    try:
        return json.loads(out[i:])
    except json.JSONDecodeError:
        # --page-all 可能拼接多个对象，取最后一个完整 JSON
        j = out.rfind('{"code"')
        return json.loads(out[j:]) if j >= 0 else {"_err": out[:400]}


def dept_name_map(workdir):
    """open_department_id -> 部门名。失败返回空 dict，不阻断主流程。

    部门树走通讯录 contact API，需 contact:department.base:readonly。bot 通常无此 scope，
    user 身份（已 auth login）一般可读，所以优先 user、回退 bot。
    """
    for ident in ("user", "bot"):
        d = api("GET", "/open-apis/contact/v3/departments/0/children",
                {"fetch_child": True, "department_id_type": "open_department_id"},
                ident=ident, page_all=True, workdir=workdir)
        if d.get("code") == 0:
            return {it["open_department_id"]: it.get("name", "") for it in d["data"]["items"]}
    sys.stderr.write(f"[warn] 部门树拉取失败，部门列将用 id：{d.get('_err') or d.get('msg')}\n")
    return {}


def val(x):
    return x.get("value") if isinstance(x, dict) else x


def name_of(x):
    return x.get("name") if isinstance(x, dict) else x


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out-dir", default="./roster")
    ap.add_argument("--status", default="2",
                    help="在职状态，逗号分隔。默认 2=在职。全部用 1,2,3,4,5")
    ap.add_argument("--xlsx-name", default=None)
    args = ap.parse_args()

    os.makedirs(args.out_dir, exist_ok=True)
    workdir = args.out_dir

    status_list = [int(s) for s in args.status.split(",") if s.strip()]
    depts = dept_name_map(workdir)

    res = api("GET", "/open-apis/ehr/v1/employees",
              {"view": "full", "status": status_list, "page_size": 100},
              ident="bot", page_all=True, workdir=workdir)
    if res.get("code") != 0:
        sys.stderr.write("拉取花名册失败。检查 bot 是否已开通 ehr:employee:readonly。\n")
        sys.stderr.write(json.dumps(res, ensure_ascii=False)[:500] + "\n")
        sys.exit(1)

    items = res["data"]["items"]
    rows = []
    for it in items:
        s = it["system_fields"]
        did = s.get("department_id")
        rows.append({
            "姓名": s.get("name"),
            "工号": s.get("employee_no"),
            "人员类型": EMP_TYPE.get(val(s.get("employee_type")), val(s.get("employee_type"))),
            "在职状态": STATUS.get(val(s.get("status")), val(s.get("status"))),
            "入职日期": s.get("hire_date"),
            "转正日期": s.get("conversion_date"),
            "试用期(月)": s.get("probation_months"),
            "部门": depts.get(did, did),
            "职务": name_of(s.get("job")),
            "职级": name_of(s.get("job_level")),
            "工作地点": name_of(s.get("work_location")),
            "直属上级": (s.get("manager") or {}).get("name") if isinstance(s.get("manager"), dict) else None,
            "手机": s.get("mobile"),
            "邮箱": s.get("email"),
        })

    # 落 JSON（含全字段原始备份）
    json.dump({"pulled_at": datetime.datetime.now().isoformat(timespec="seconds"),
               "count": len(rows), "rows": rows},
              open(os.path.join(args.out_dir, "roster.json"), "w", encoding="utf-8"),
              ensure_ascii=False, indent=1)

    # 落 Excel
    xlsx = args.xlsx_name or f"花名册_{datetime.date.today().isoformat()}.xlsx"
    xlsx_path = os.path.join(args.out_dir, xlsx)
    write_xlsx(rows, xlsx_path)

    print(f"OK: {len(rows)} 人  |  Excel: {xlsx_path}  |  JSON: {os.path.join(args.out_dir,'roster.json')}")
    from collections import Counter
    c = Counter(r["人员类型"] for r in rows)
    print("人员类型分布:", dict(c))


def write_xlsx(rows, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "花名册"
    cols = list(rows[0].keys()) if rows else []
    ws.append(cols)
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for c in ws[1]:
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    intern_fill = PatternFill("solid", fgColor="FFF2CC")
    for r in rows:
        ws.append([r.get(k) for k in cols])
        if r.get("人员类型") == "实习":
            for c in ws[ws.max_row]:
                c.fill = intern_fill
        for c in ws[ws.max_row]:
            c.border = border
    widths = {"姓名": 10, "工号": 9, "人员类型": 9, "在职状态": 9, "入职日期": 12,
              "转正日期": 12, "试用期(月)": 9, "部门": 16, "职务": 22, "职级": 10,
              "工作地点": 10, "直属上级": 10, "手机": 16, "邮箱": 24}
    for i, k in enumerate(cols, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = widths.get(k, 12)
    ws.freeze_panes = "A2"
    wb.save(path)


if __name__ == "__main__":
    main()
