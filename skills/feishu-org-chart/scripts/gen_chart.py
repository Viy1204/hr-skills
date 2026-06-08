# -*- coding: utf-8 -*-
"""
把已拉取的飞书部门树 + 成员数据生成为画板 DSL（diagram.json）。

输入（--data-dir 目录下，由 pull_org.py 产出）：
  departments_raw.json   contact/v3 .../children 的原始响应（含 data.items）
  members/<open_department_id>.json   各部门 find_by_department 的原始响应

岗位（可选，--roster 指定花名册 xlsx）：飞书 find_by_department 通常不返回 job_title，
所以岗位按「姓名」从花名册匹配。没有花名册时节点只显示姓名。

输出：dagre（rankdir TB）自动布局，每个部门一张 flex 卡片（部门名+人数，
成员逐行「· 姓名  岗位」，负责人标注），部门间按 parent_department_id 连父子。
这样能容纳几十个部门、上百人，不受手工树「≤5 子节点」的限制。

用法：
  python gen_chart.py --data-dir ./org-data --company-name "公司名" \
      [--roster roster.xlsx --name-col 1 --title-col 4] --out diagram.json
"""
import argparse, glob, json, os, sys

sys.stdout.reconfigure(encoding="utf-8")

# 按层级递进的配色：root 深 → L1 蓝 → L2 绿 → L3 紫 → L4 橙
TIER = {
    1: ("#E1EDFF", "#4E83FD"),
    2: ("#E4FBE6", "#34C724"),
    3: ("#F0EAFF", "#7C5CFC"),
    4: ("#FFEFD9", "#FF8800"),
}


def load_roster(path, name_col, title_col):
    if not path:
        return {}
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    name2title = {}
    for sh in wb.sheetnames:
        for r in wb[sh].iter_rows(min_row=1, values_only=True):
            if name_col >= len(r):
                continue
            nm = r[name_col]
            title = r[title_col] if title_col < len(r) else None
            if nm and str(nm).strip():
                nm = str(nm).strip()
                if nm not in name2title and title:
                    name2title[nm] = str(title).strip()
    return name2title


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--data-dir", required=True)
    ap.add_argument("--company-name", default="组织架构")
    ap.add_argument("--roster", default=None, help="花名册 xlsx（可选，按姓名补岗位）")
    ap.add_argument("--name-col", type=int, default=1, help="花名册姓名列(0基)")
    ap.add_argument("--title-col", type=int, default=4, help="花名册岗位列(0基)")
    ap.add_argument("--out", default="diagram.json")
    args = ap.parse_args()

    dd = args.data_dir
    depts = json.load(open(os.path.join(dd, "departments_raw.json"), encoding="utf-8"))["data"]["items"]
    dept_by_id = {d["open_department_id"]: d for d in depts}

    members, openid2name = {}, {}
    for f in glob.glob(os.path.join(dd, "members", "*.json")):
        did = os.path.basename(f)[:-5]
        try:
            its = json.load(open(f, encoding="utf-8")).get("data", {}).get("items", []) or []
        except Exception:
            its = []
        members[did] = its
        for u in its:
            openid2name[u["open_id"]] = u["name"]

    name2title = load_roster(args.roster, args.name_col, args.title_col)
    matched = sum(1 for n in openid2name.values() if name2title.get(n))

    def depth(did, seen=None):
        seen = seen or set()
        p = dept_by_id[did].get("parent_department_id", "0")
        if p == "0" or p not in dept_by_id or did in seen:
            return 1
        return 1 + depth(p, seen | {did})

    def card(did):
        d = dept_by_id[did]
        fill, border = TIER[min(depth(did), 4)]
        leader = d.get("leader_user_id")
        mems = members.get(did, [])
        ordered = sorted(mems, key=lambda u: (u["open_id"] != leader, u["name"]))
        lines = []
        for u in ordered:
            t = name2title.get(u["name"], "")
            tag = " （负责人）" if u["open_id"] == leader else ""
            lines.append(f"· {u['name']}" + (f"  {t}" if t else "") + tag)
        if not lines:
            ln = openid2name.get(leader)
            lines = [f"· {ln}  （负责人）"] if ln else ["（无直属成员）"]
        return {
            "type": "frame", "id": did, "layout": "vertical",
            "gap": 6, "padding": [10, 12], "width": 240, "height": "fit-content",
            "fillColor": fill, "borderColor": border, "borderWidth": 2,
            "borderRadius": 8, "alignItems": "stretch",
            "children": [
                {"type": "text", "width": "fill-container", "height": "fit-content",
                 "text": [{"content": f"{d['name']}（{len(mems)}人）", "bold": True,
                           "fontSize": 15, "color": "#1F2329"}], "textAlign": "center"},
                {"type": "text", "width": "fill-container", "height": "fit-content",
                 "text": "\n".join(lines), "fontSize": 12,
                 "textColor": "#3F454D", "textAlign": "left"},
            ],
        }

    children = [{
        "type": "rect", "id": "root", "width": 300, "height": "fit-content",
        "fillColor": "#1F2329", "borderColor": "#1F2329", "borderWidth": 2, "borderRadius": 10,
        "text": [{"content": f"{args.company_name}\n", "bold": True, "fontSize": 22, "color": "#FFFFFF"},
                 {"content": f"组织架构 · 共 {len(openid2name)} 人 / {len(depts)} 部门",
                  "fontSize": 13, "color": "#FFFFFF"}],
        "textAlign": "center", "verticalAlign": "middle",
    }]
    for did in dept_by_id:
        children.append(card(did))

    edges = []
    for did, d in dept_by_id.items():
        p = d.get("parent_department_id", "0")
        edges.append(["root", did] if (p == "0" or p not in dept_by_id) else [p, did])

    doc = {"version": 2, "nodes": [{
        "type": "frame", "id": "org", "layout": "dagre",
        "width": "fit-content", "height": "fit-content", "padding": 60, "gap": 0,
        "fillColor": "#FFFFFF", "borderColor": "#FFFFFF", "borderWidth": 1,
        "layoutOptions": {"rankdir": "TB", "nodesep": 30, "ranksep": 80, "edges": edges},
        "children": children,
    }]}

    json.dump(doc, open(args.out, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    print(f"depts={len(depts)} people={len(openid2name)} title_matched={matched} -> {args.out}")


if __name__ == "__main__":
    main()
